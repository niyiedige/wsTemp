import openpyxl

import datetime as DT
# reinstall openpxyl if anythhing happen
# to modifide the output :1 change the title here 2.change the source of the output in getinfo
def output(output):
    today = DT.date.today()
    threeday = today - DT.timedelta(days=3)
    title = ['link', 'company', 'CM','id','title', 'department', 'location', 'date','time type']
    wb=openpyxl.load_workbook('result.xlsx')
    try :
        #depend on the frequency
        ws1=wb.get_sheet_by_name(str(threeday))
        wb.remove_sheet(ws1)
    except KeyError:
        pass
    try :
        #depend on the frequency
        ws1=wb.get_sheet_by_name(str(today))
        wb.remove_sheet(ws1)
    except KeyError:
        pass
    ws=wb.create_sheet(str(today))
    for col in ws.iter_cols(min_row=1, max_col=8):
        for cell in col:
            cell.value=title.pop()
    rownum=1

    while output:
        colnum = 0
        rownum=rownum+1
    #    print(output)
        job=output.pop()
    #    print(1)
        while job:
            colnum+=1
            a = job.pop()
            ws.cell(row=rownum,column=colnum).value = a



    wb.save("result.xlsx")
